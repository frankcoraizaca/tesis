from django.core.exceptions import RequestAborted
from django.db.models import query
from django.db.models.query import QuerySet
from django.forms.models import fields_for_model
from django.http import request
from django.shortcuts import render, redirect
from django.urls.base import is_valid_path
from django.views.generic import (TemplateView, 
ListView, UpdateView,DeleteView,
CreateView
)
from .models import Programa
from openpyxl import Workbook
from  applications.areas.models import areaslista
from django.http import HttpResponse, Http404
from .forms import ProgramaForm

from django.http import StreamingHttpResponse
 
from django.core.files import File
from django.urls import reverse_lazy
import pandas as pd
from pandas import ExcelWriter
import openpyxl
import numpy as np 
import os.path
from tablib import Dataset
os.environ.setdefault("DJANGO_SETTINGS","empledo.settings")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "StreamingHttpResponse.settings")
 

# Create your views here.

def programa_edit(request,id_Programa):
    programa=Programa.objects.get(id=id_Programa)
    if request.method== 'GET':
        form = ProgramaForm(instance=programa)
    else:
        form= ProgramaForm(request.POST,instance=programa)
        if form.is_valid():
            form.save()
        return redirect('ventilcion:listaprograma')
    return render(request, 'ventilcion/programaventi.html', {'form':form})


class registroventi(CreateView):
    model = Programa
    template_name = "ventilcion/programaventi.html"
    form_class= ProgramaForm
    success_url = '.'
  

class listaListView(ListView):
    model = Programa
    template_name = "ventilcion/listaventi.html"

    
class ProgramadateView(UpdateView):
    template_name = "ventilcion/editar.html"
    model = Programa
    form_class=ProgramaForm
    success_url='.'


class ProgramaDeleteView(DeleteView):
    model = Programa
    template_name = "ventilcion/eliminard.html"
    success_url= reverse_lazy('listaprograma')


def suma(request):
    qs=Programa.objects.all().values()
    data = pd.DataFrame(qs)
    data = data.fillna(0)
    data = data.fillna({'de1':'x'})
    qs1=areaslista.objects.all().values()
    data1 = pd.DataFrame(qs1)
   




    data ['flujo']=0
    data ['D.equivalente']=0
    data ['lado1']=0
    data ['lado2']=0
    data ['velocidad']=0
    data ['D.real']=0
    data ['Perdidas por longitud']=0
    data ['Perdidas acumulada']=0
    data ['Recinto']=0
    qs2=[0]
    data2=pd.DataFrame(qs2,columns=['d.principal'])
    qs3=[0]
    data3=pd.DataFrame(qs3,columns=['Caída de presión  del sistema de ventilación [Pa]'])
    qs4=[0]
    data4=pd.DataFrame(qs4,columns=['Pérdidas por longitud [Pa/m]'])
    qs5=[0]
    data5=pd.DataFrame(qs5,columns=['Tramo'])


    


    

   #llenado la columna de flujo 
   # flujo requerido por cada region
    for i in range(0,len(data)):
        if data.loc[i,'Dfinal']==True:
            for x in range(0,len(data1)):
                if data.loc[i,'tipoarea_id']>=1 and data.loc[i,'tipoarea_id']<=48 :
                    if data.loc[i,'tipoarea_id']== data1.loc[x,'id']:
                        resul=data.loc[i,'Darea']*data1.loc[x,'litrosporarea']+data.loc[i,'npersonas']*data1.loc[x,'listrosporpersona']
                        data.loc[i,'flujo']=resul
                        data.loc[i,'Recinto']=data1.loc[x,'areas']
                        
                        
                elif data.loc[i,'tipoarea_id']>=56:
                    if data.loc[i,'tipoarea_id']== data1.loc[x,'id']:
                        resul=(data.loc[i,'Darea']*data1.loc[x,'litrosporarea']*3.5)/3.6
                        data.loc[i,'flujo']=resul
                        data.loc[i,'Recinto']=data1.loc[x,'areas']



                
                    
        else:
            resul=0
            data.loc[i,'flujo']=resul
    
   #flujo por cada tramo 
    r=False   
    while r==False:
        r=True
        for y in range(0,len(data)):
            if data.loc[y,'flujo']==0:
                if data.loc[y,'de3']=='':
                    for l in range(0,len(data)):
                        if data.loc[y,'de1']==data.loc[l,'Tramo'] and data.loc[l,'flujo']!=0:
                            if data.loc[y,'Unión_id']!=4:
                                for p in range(0,len(data)):
                                    if data.loc[y,'de2']==data.loc[p,'Tramo'] and data.loc[p,'flujo']!=0:
                                         data.loc[y,'flujo']=data.loc[l,'flujo']+data.loc[p,'flujo']
                            elif data.loc[y,'Unión_id']==4:
                                for x in range(0,len(data1)):
                                    if data.loc[y,'tipoarea_id']>=1 and data.loc[y,'tipoarea_id']<=48 :
                                        if data.loc[y,'tipoarea_id']== data1.loc[x,'id']:
                                            resul=data.loc[y,'Darea']*data1.loc[x,'litrosporarea']+data.loc[y,'npersonas']*data1.loc[x,'listrosporpersona']
                                            data.loc[y,'Recinto']=data1.loc[x,'areas']
                                    elif data.loc[y,'tipoarea_id']>=56:
                                        if data.loc[y,'tipoarea_id']== data1.loc[x,'id']:
                                           resul=(data.loc[y,'Darea']*data1.loc[x,'litrosporarea']*3.5)/3.6
                                           data.loc[y,'Recinto']=data1.loc[x,'areas']
                                data.loc[y,'flujo']=data.loc[l,'flujo']+resul

                if data.loc[y,'de3']!='':
                    for l in range(0,len(data)):
                        if data.loc[y,'de1']==data.loc[l,'Tramo'] and data.loc[l,'flujo']!=0:
                            for p in range(0,len(data)):
                                if data.loc[y,'de2']==data.loc[p,'Tramo'] and data.loc[p,'flujo']!=0:
                                   for w in range(0,len(data)):
                                       if data.loc[y,'de3']==data.loc[w,'Tramo'] and data.loc[w,'flujo']!=0:
                                          data.loc[y,'flujo']=data.loc[l,'flujo']+data.loc[p,'flujo']+data.loc[w,'flujo']

        for y in range(0,len(data)):
            if data.loc[y,'flujo']==0:
                r=False
    
    

    # diámetro equivalente 
   
    ve=float(request.POST.get('num1'))

   
    
    Dh = np.sqrt(((data["flujo"].max()/1000)/ve)*(4/3.1415))
    de=1.2 #densidad del aire
    Re=(de*Dh*ve)/0.000018 #Reynolds
    e1=0.1219     #rugosidad
    Rr=e1/(Dh*1000)     #rugosidad relativa
    REs= friccion(Re,Rr)    #factor de fricion 
    Pl=REs*((de*np.square(ve))/(2*Dh))
 
    #diametro equivalente 
    for i in range(0,len(data)):
        Y=de*0.8105*np.square((data.loc[i,'flujo']/1000))
        P=(de*4*(data.loc[i,'flujo']/1000))/(0.000024*3.1415)
        Dequi=Dequivalente(Y,P,Pl,e1)
        data.loc[i,'D.equivalente']=Dequi*1000 #transformar a mm

    #selecionar diametros  y velocidad real
    for i in range(0,len(data)):
        x1=data.loc[i,'D.equivalente']
        if x1<(ductos['du'][0]):
            data.loc[i,'lado1']=(ductos['lmayor'][0])
            data.loc[i,'lado2']=(ductos['lmenor'][0])
            data.loc[i,'velocidad']=((data.loc[i,'flujo'])/1000)/(ductos['area'][0])
            data.loc [i,'D.real']=ductos['du'][0]
        else:
            for y in range(0,26):
                x2=(ductos['du'][y])
                x3=(ductos['du'][y+1])
                if x1 >= x2 and x1<x3:
                    x4=((x3-x2)/2)+x2
                   
                    if x1 < x4:
                        #print(y)
                        data.loc[i,'lado1']=(ductos['lmayor'][y])
                        data.loc[i,'lado2']=(ductos['lmenor'][y])
                        velo=(ductos['area'][y])
                        data.loc[i,'velocidad']=((data.loc[i,'flujo'])/1000)/velo
                        data.loc [i,'D.real']=ductos['du'][y]
                    elif x1 > x4:
                        #print(y+1)
                        data.loc[i,'lado1']=(ductos['lmayor'][y+1])
                        data.loc[i,'lado2']=(ductos['lmenor'][y+1])
                        velo=(ductos['area'][y+1])
                        data.loc[i,'velocidad']=((data.loc[i,'flujo'])/1000)/velo
                        data.loc [i,'D.real']=ductos['du'][y+1]
    
    #perdidas por longuitud.
 
    for i in range(0,len(data)):
        Rrre=e1/(data.loc [i,'D.real'])
        Rere=(de*((data.loc [i,'D.real'])/1000)*data.loc[i,'velocidad'])/0.000024
        Resre=friccion(Rere,Rrre)

        Plre=Resre*(de*np.square(data.loc[i,'velocidad']))/(2*(data.loc [i,'D.real'])/1000)
        data.loc[i,'Factor de pérdida']=Plre
        #perdidas por codos.
        Pc=7*data.loc[i,'lado2']*data.loc[i,'Ntramo']/1000
        Plre=Plre*(data.loc[i,'d1distancia']+Pc)
        data.loc[i,'Perdidas por longitud']=Plre
   
 #perdida acumulada 
    for i in range(0,len(data)):
        if data.loc[i,'Dfinal']==True:
            data.loc[i,'Perdidas acumulada']=data.loc[i,'Perdidas por longitud']
            
        else:
            resul=0
            data.loc[i,'Perdidas acumulada']=resul

    
    r=False   
    while r==False:
        r=True
        for y in range(0,len(data)):
            if data.loc[y,'Perdidas acumulada']==0:
               
                if data.loc[y,'de3']=='':
                    for l in range(0,len(data)):
                        
                        if data.loc[y,'de1']==data.loc[l,'Tramo'] and data.loc[l,'Perdidas acumulada']!=0:
                            if data.loc[y,'Unión_id']!=4:

                                for p in range(0,len(data)):
                                   if data.loc[y,'de2']==data.loc[p,'Tramo'] and data.loc[p,'Perdidas acumulada']!=0:
                                        
                                        data.loc[y,'Perdidas acumulada']=data.loc[l,'Perdidas acumulada']+data.loc[p,'Perdidas acumulada']+data.loc[y,'Perdidas por longitud']
                            elif data.loc[y,'Unión_id']==4:
                                
                                data.loc[y,'Perdidas acumulada']=data.loc[l,'Perdidas acumulada']+data.loc[y,'Perdidas por longitud']
                        
                if data.loc[y,'de3']!='':
                    for l in range(0,len(data)):

                        if data.loc[y,'de1']==data.loc[l,'Tramo'] and data.loc[l,'Perdidas acumulada']!=0:
                            for p in range(0,len(data)):
                                if data.loc[y,'de2']==data.loc[p,'Tramo'] and data.loc[p,'Perdidas acumulada']!=0:
                                   for w in range(0,len(data)):
                                       if data.loc[y,'de3']==data.loc[w,'Tramo'] and data.loc[w,'Perdidas acumulada']!=0:
                                          data.loc[y,'Perdidas acumulada']=data.loc[l,'Perdidas acumulada']+data.loc[p,'Perdidas acumulada']+data.loc[w,'Perdidas acumulada']+data.loc[y,'Perdidas por longitud']
        for y in range(0,len(data)):
            if data.loc[y,'Perdidas acumulada']==0:
                r=False
 
#identificar el conducto principal
    maxClm = data['Perdidas acumulada'].idxmax() 
    tr1=data.loc[maxClm ,'de1']
    tr2=data.loc[maxClm ,'de2']
    tr3=data.loc[maxClm ,'de3']
    data2.loc[0,'d.principal']=data.loc [maxClm,'Tramo']
    data2.loc[0,'lado1']=data.loc [maxClm,'lado1']
    data2.loc[0,'lado2']=data.loc [maxClm,'lado2']
    data2.loc[0,'Distancia [m]']=data.loc [maxClm,'d1distancia']
    data2.loc[0,'Unión_id']=data.loc [maxClm,'Unión_id']
    data2.loc[0,'Ntramo']=data.loc [maxClm,'Ntramo']
    data2.loc[0,'flujo']=data.loc [maxClm,'flujo']
    data2.loc[0,'velocidad']=data.loc [maxClm,'velocidad']
    data2.loc[0,'Factor de pérdida']=data.loc[maxClm,'Factor de pérdida']
    data2.loc[0,'Perdidas por longitud']=data.loc [maxClm,'Perdidas por longitud']
    p=True
    o=0

  
    while p==True:
    
        i3=-1
        i2=-1
   
        for i in range(0,len(data)):
            if data.loc[i,'Tramo']==tr1:
                i1=i
            if data.loc[i,'Tramo']==tr2:
                i2=i
            if data.loc[i,'Tramo']==tr3:
                i3=i


        if i2<=0:
            o=1+o
            data2.loc[o,'d.principal']=data.loc [i1,'Tramo']
            data2.loc[o,'lado1']=data.loc [i1,'lado1']
            data2.loc[o,'lado2']=data.loc [i1,'lado2']
            data2.loc[o,'Distancia [m]']=data.loc [i1,'d1distancia']
            data2.loc[o,'Unión_id']=data.loc [i1,'Unión_id']
            data2.loc[o,'Ntramo']=data.loc [i1,'Ntramo']
            data2.loc[o,'flujo']=data.loc [i1,'flujo']
            data2.loc[o,'velocidad']=data.loc [i1,'velocidad']
            data2.loc[o,'Factor de pérdida']=data.loc[i1,'Factor de pérdida']
            data2.loc[o,'Perdidas por longitud']=data.loc [i1,'Perdidas por longitud']
            tr1=data.loc[i1 ,'de1']
            tr2=data.loc[i1 ,'de2']
            tr3=data.loc[i1 ,'de3']
        elif i3 >=0:
           o=1+o

           if data.loc[i1,'Perdidas acumulada']>=data.loc[i2,'Perdidas acumulada'] and data.loc[i1,'Perdidas acumulada']>=data.loc[i3,'Perdidas acumulada']:
        
                data2.loc[o,'d.principal']=data.loc [i1,'Tramo']
                data2.loc[o,'lado1']=data.loc [i1,'lado1']
                data2.loc[o,'lado2']=data.loc [i1,'lado2']
                data2.loc[o,'Distancia [m]']=data.loc [i1,'d1distancia']
                data2.loc[o,'Unión_id']=data.loc [i1,'Unión_id']
                data2.loc[o,'Ntramo']=data.loc [i1,'Ntramo']
                data2.loc[o,'flujo']=data.loc [i1,'flujo']
                data2.loc[o,'velocidad']=data.loc [i1,'velocidad']
                data2.loc[o,'Factor de pérdida']=data.loc[i1,'Factor de pérdida']
                data2.loc[o,'Perdidas por longitud']=data.loc [i1,'Perdidas por longitud']
                tr1=data.loc[i1 ,'de1']
                tr2=data.loc[i1 ,'de2']
                tr3=data.loc[i1 ,'de3']
              

           elif data.loc[i2,'Perdidas acumulada']>=data.loc[i1,'Perdidas acumulada'] and data.loc[i2,'Perdidas acumulada']>=data.loc[i3,'Perdidas acumulada']:
                data2.loc[o,'d.principal']=data.loc [i2,'Tramo']
                data2.loc[o,'lado1']=data.loc [i2,'lado1']
                data2.loc[o,'lado2']=data.loc [i2,'lado2']
                data2.loc[o,'Distancia [m]']=data.loc [i2,'d1distancia']
                data2.loc[o,'Unión_id']=data.loc [i2,'Unión_id']
                data2.loc[o,'Ntramo']=data.loc [i2,'Ntramo']
                data2.loc[o,'flujo']=data.loc [i2,'flujo']
                data2.loc[o,'velocidad']=data.loc [i2,'velocidad']
                data2.loc[o,'Factor de pérdida']=data.loc[i2,'Factor de pérdida']
                data2.loc[o,'Perdidas por longitud']=data.loc [i2,'Perdidas por longitud']
                tr1=data.loc[i2 ,'de1']
                tr2=data.loc[i2 ,'de2']
                tr3=data.loc[i2 ,'de3']
               
           else:
                data2.loc[o,'d.principal']=data.loc [i3,'Tramo']
                data2.loc[o,'lado1']=data.loc [i3,'lado1']
                data2.loc[o,'lado2']=data.loc [i3,'lado2']
                data2.loc[o,'Distancia [m]']=data.loc [i3,'d1distancia']
                data2.loc[o,'Unión_id']=data.loc [i3,'Unión_id']
                data2.loc[o,'Ntramo']=data.loc [i3,'Ntramo']
                data2.loc[o,'flujo']=data.loc [i3,'flujo']
                data2.loc[o,'velocidad']=data.loc [i3,'velocidad']
                data2.loc[o,'Factor de pérdida']=data.loc[i3,'Factor de pérdida']
                data2.loc[o,'Perdidas por longitud']=data.loc [i3,'Perdidas por longitud']
                tr1=data.loc[i3 ,'de1']
                tr2=data.loc[i3 ,'de2']
                tr3=data.loc[i3 ,'de3']
              
        else:
           o=1+o
           
           if data.loc[i1,'Perdidas acumulada']>=data.loc[i2,'Perdidas acumulada'] :
                data2.loc[o,'d.principal']=data.loc [i1,'Tramo']
                data2.loc[o,'lado1']=data.loc [i1,'lado1']
                data2.loc[o,'lado2']=data.loc [i1,'lado2']
                data2.loc[o,'Distancia [m]']=data.loc [i1,'d1distancia']
                data2.loc[o,'Unión_id']=data.loc [i1,'Unión_id']
                data2.loc[o,'Ntramo']=data.loc [i1,'Ntramo']
                data2.loc[o,'flujo']=data.loc [i1,'flujo']
                data2.loc[o,'velocidad']=data.loc [i1,'velocidad']
                data2.loc[o,'Factor de pérdida']=data.loc[i1,'Factor de pérdida']
                data2.loc[o,'Perdidas por longitud']=data.loc [i1,'Perdidas por longitud']
                tr1=data.loc[i1 ,'de1']
                tr2=data.loc[i1 ,'de2']
                tr3=data.loc[i1 ,'de3']
               
           else :
                data2.loc[o,'d.principal']=data.loc [i2,'Tramo']
                data2.loc[o,'lado1']=data.loc [i2,'lado1']
                data2.loc[o,'lado2']=data.loc [i2,'lado2']
                data2.loc[o,'Distancia [m]']=data.loc [i2,'d1distancia']
                data2.loc[o,'Unión_id']=data.loc [i2,'Unión_id']
                data2.loc[o,'Ntramo']=data.loc [i2,'Ntramo']
                data2.loc[o,'flujo']=data.loc [i2,'flujo']
                data2.loc[o,'velocidad']=data.loc [i2,'velocidad']
                data2.loc[o,'Factor de pérdida']=data.loc[i2,'Factor de pérdida']
                data2.loc[o,'Perdidas por longitud']=data.loc [i2,'Perdidas por longitud']
                tr1=data.loc[i2 ,'de1']
                tr2=data.loc[i2 ,'de2']
                tr3=data.loc[i2 ,'de3']
            
            

           
        if tr1=='':
            p=False
    
    

    data2 ['Perdidas en accesorios']=0  
    
    for z in range(0,len(data2)-1):
        tt1=data2.loc[z,'lado1']
        tt2=data2.loc[z+1,'lado1']
        if tt1==tt2:
            if data2.loc[z,'Unión_id']==2:
                data2.loc[z,'Perdidas en accesorios']=(7*data2.loc[z,'lado2']/1000)*Pl
            
        elif data2.loc[z,'Unión_id']==2:
            data2.loc[z,'Perdidas en accesorios']=(7*data2.loc[z,'lado2']/1000)*Pl

        elif data2.loc[z,'Unión_id']==3 or data2.loc[z,'Unión_id']==4 or data2.loc[z,'Unión_id']==1:
            data2.loc[z,'Perdidas en accesorios']=abs((0.311*(((data2.loc[z,'velocidad']/4)**2)-((data2.loc[z+1,'velocidad']/4)**2)))*9.8)

    
    
    #ganancias 
    data2 ['ganancias']=0
    data2['P.totales']=0
    Ga=((((((data2['velocidad'].iloc[0])**2)/1.278)-(((data2['velocidad'].iloc[-1])**2)/1.278))*0.75))
    
    #perdidas totales
    for i in range(0,len(data2)):
        data2.loc[i,'P.totales']=data2.loc[i,'Perdidas por longitud']
    
    Total=0 
    ACC=0 
    for i in range(0,len(data2)):
        if data2.loc[i,'P.totales']>0:
            Total=Total+data2.loc[i,'P.totales']
            ACC=ACC+data2.loc[i,'Perdidas en accesorios']
   
    F1=data2['d.principal'].iloc[-1]

    E=0.0625  #area del difusor 
    
    for i in range(0,len(data2)):
        if F1==data2.loc[i,'d.principal']:
            difu=((data2.loc[i,'flujo']/E)**1.928)*0.00000007
            

    data3.loc[0,'Flujo de aire requerido [l/s]']="{:.2f}".format(data.loc [maxClm,'flujo'] )
    data3.loc[0,'Pérdidas por longitud [Pa]']="{:.2f}".format(Total)    
    data3.loc[0,'Pérdidas por accesorios [Pa]']="{:.2f}".format(ACC )     
    data3.loc[0,'Caída de presión  del sistema de ventilación [Pa]']= "{:.2f}".format(Total+9-Ga+ACC)   
    data3.loc[0,'Caída de presión por el difusor [Pa]']=9
    data3.loc[0,'Recuperación estática[Pa]']="{:.2f}".format(Ga)   
    
    
    data4.loc[0,'Pérdidas por longitud [Pa/m]' ]=Pl

    data.rename(columns={'de1':'Derivación 1',
                          'de2':'Derivación 2',
                          'de3':'Derivación 3',
                          'd1distancia':'Distancia [m]',
                          'flujo':'Flujo [L/s]',
                          'D.real':'Diámetro [mm]',
                          'tipoarea_id':'TP',
                          'lado1':'Lado Mayor [mm]',
                          'lado2':'Lado Menor [mm]',
                          'Codigo':'Código',
                          'Darea':'Área',
                          'velocidad':'Velocidad [m/s]',
                          'Perdidas por longitud':'Pérdidas por longitud [Pa]',
                        'de3':'Derivación 3'},

               inplace=True)

    
    data2.rename(columns={'d.principal':'Tramo',
                          'lado1':'Lado Mayor [mm]',
                          'lado2':'Lado Menor [mm]',
                          'flujo':'Flujo [L/s]',
                          'velocidad':'Velocidad [m/s]',
                          'Perdidas por longitud':'Pérdidas por longitud [Pa]',
                          'Perdidas en accesorios':'Pérdidas en accesorios [Pa]',
                          'ganancias':'Ganancias estática [Pa]',
                          'P.totales':'Pérdidas de presión total [Pa]'},
               inplace=True)

   

  
    data['Flujo [L/s]']=pd.Series([round(val, 2) for val in data['Flujo [L/s]']])
    data['Velocidad [m/s]']=pd.Series([round(val, 2) for val in data['Velocidad [m/s]']])
    data['Pérdidas por longitud [Pa]']=pd.Series([round(val, 3) for val in data['Pérdidas por longitud [Pa]']])

      
    data2['Pérdidas en accesorios [Pa]']=pd.Series([round(val, 3) for val in data2['Pérdidas en accesorios [Pa]']])
    data2['Velocidad [m/s]']=pd.Series([round(val, 2) for val in data2['Velocidad [m/s]']])


    
    qs5=data[['Tramo','Derivación 1','Derivación 2','Derivación 3','Distancia [m]','Flujo [L/s]','Velocidad [m/s]','Pérdidas por longitud [Pa]']]
    data5 = pd.DataFrame(qs5)  
    qs7=data[['Tramo','Lado Mayor [mm]','Lado Menor [mm]']]
    data7 = pd.DataFrame(qs7)
    ### manejo de tablas
    #qs6=data2[['Tramo','Lado Mayor [mm]','Lado Menor [mm]','Distancia [m]','Velocidad [m/s]','Pérdidas por longitud [Pa]','Pérdidas en accesorios [Pa]','Ganancias estática [Pa]','Pérdidas de presión total [Pa]']]
    qs6=data2[['Tramo','Distancia [m]','Lado Mayor [mm]','Lado Menor [mm]','Velocidad [m/s]','Pérdidas en accesorios [Pa]']]

    data6 = pd.DataFrame((qs6))
    
    context = {
        'df': data.to_html(),'df1': data1.to_html(),'df2': data2.to_html(),'df3': data3.to_html(),'df4': data4.to_html(),'df5': data5.to_html(),
        'df6': data6.to_html(),'df7': data7.to_html(),
    }
    return render(request,'ventilcion/pruebaf.html',context)




###ecuacion white
def friccion(Ree,Rrr):
    f=1
    Rrr1=Rrr/3.71
    Ree1=2.51/Ree
    c=0.05
    while abs(f)>0.0001:
        B=1/np.square((-2*np.log10((Rrr1)+(Ree1/(np.sqrt(c))))))
        f=B-c
        c=B
    return c
###ecuación white para determinar el diametro equivalente 
def Dequivalente(Y1,P1,Pl1,e1):
    A1=(Y1/Pl1)**(1/5)
    A2=e1/3710
    A3=2.51/P1
    A4=Pl1/Y1
    d=0.5
    f1=1
    while abs(f1)>0.001:
        B=A1/(-2*np.log10((A2/d)+((A3*d)/((A4*(d**5))**(1/2)))))**0.4
        f1=B-d
        d=B
    return d

ductos={
    'area':[0.0138,0.016,0.0182,0.028,0.0375,0.0468,0.0587,0.0702,0.0844,0.0984,0.115,0.1312,0.1502,0.1688,0.1901,0.2110,0.2346,0.2579,0.2838,0.3095,0.3378,0.3658,0.3965,0.4269,0.4598,0.4926,0.5280,0.5634,0.6013],
    'lmenor':[100,  100,    100,150  ,200,   200,      250,   250,   300,   300,  350,   350,   400,   400,   450,   450,   500,   500,   550,   550,   600,   600,   650,   650,700,700,750,750,800],
    'lmayor':[150,  175,    200,200  ,200,   250,      250,   300,   300,   350,  350,   400,   400,   450,   450,   500,   500,   550,   550,   600,   600,   650,   650,   700,700,750,750,800,800],
    'du':[133,     143,   152.3,188.9,218.6,244.1,   273.3, 299.1, 327.9,   354,382.6, 408.8, 437.3, 463.6, 491.9, 518.4, 546.6, 573.1, 601.2, 627.8, 655.9, 682.5, 710.6,  737.3,765.2,792,820,847,875]}



def informe(request):
    eliminarprograma=Programa.objects.all()
    
    eliminarprograma.delete()
    return render(request,'ventilcion/informe.html')


def Velocidad(request):
    ve=request.POST.get('num1')
    d1=request.POST.get('df')
    qs=Programa.objects.all().values()
    data = pd.DataFrame(qs)
    ###validaación de valor único 
    """for i in range(0,len(data)-1):
        p=True
        c=i+1
        y=0
        while p==True:
            c=c+y
            if data.loc[i,'Tramo']==data.loc[c,'Tramo']:
                print(data.loc[i,'Tramo'],data.loc[c,'Tramo'])
                
                print('error')
            if c>=95:
                p=False 
            y=1 """
    #validacion de tramo final 
    
    return render(request,'ventilcion/velocidad.html',{'ve':ve})

def flujorequerido(request):
 
    return render(request,'ventilcion/flujo.html')


def guardar(request):
    
    if request.method =="POST":
        tex2=str(request.POST.get('text1'))
        x1='.xlsx'
        x1=(tex2+str(x1))
        personas = Programa.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws['B1'] = 'id'
        ws['C1'] = 'Tramo'
        ws['D1'] = 'Derivación final'
        ws['E1'] = 'Dericación 1'
        ws['F1'] = 'Dericación 2'  
        ws['G1'] = 'Dericación 3' 
        ws['H1'] = 'Distancia' 
        ws['I1'] = 'Número de codos' 
        ws['J1'] = 'Tipo de area' 
        ws['k1'] = 'Código' 
        ws['L1'] = 'Área' 
        ws['M1'] = 'Número de personas' 
        ws['N1'] = 'Tipo de unión' 

        cont=2
        o=0
    
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            o=o+1
            ws.cell(row=cont,column=2).value = o
            ws.cell(row=cont,column=3).value = persona.Tramo
            
            ws.cell(row=cont,column=4).value = persona.Dfinal
            ws.cell(row=cont,column=5).value = persona.de1
            ws.cell(row=cont,column=6).value = persona.de2
            ws.cell(row=cont,column=7).value = persona.de3
            ws.cell(row=cont,column=8).value = persona.d1distancia
            ws.cell(row=cont,column=9).value = persona.Ntramo
            ws.cell(row=cont,column=10).value = persona.tipoarea_id
            ws.cell(row=cont,column=11).value = persona.Codigo
            ws.cell(row=cont,column=12).value = persona.Darea
            ws.cell(row=cont,column=13).value = persona.npersonas
            ws.cell(row=cont,column=14).value = persona.Unión_id
            
            cont = cont + 1
        nombre_archivo =x1
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
    return render(request,'ventilcion/guardar.html')

def cargar(request):
    if request.method =="POST":
       nuevas_personas = request.FILES['file']
       wb = openpyxl.load_workbook(nuevas_personas)
       worksheet = wb["Sheet"]
       

       excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

       for row in worksheet.iter_rows():
            
            row_data = list()
            for cell in row:
              row_data.append(str(cell.value))
            excel_data.append(row_data)
            dataD = pd.DataFrame(excel_data)
       
      
       for i in range(1,len(dataD)):
       #for i in range(1,5):
           
           a=dataD.loc[i,2]
           b=dataD.loc[i,3]
           c=dataD.loc[i,4]
           d=dataD.loc[i,5]
           e=dataD.loc[i,6]
           f=dataD.loc[i,7]
           g=dataD.loc[i,8]
           h=dataD.loc[i,9]
           m=dataD.loc[i,10]
           j=dataD.loc[i,11]
           k=dataD.loc[i,12]
           l=dataD.loc[i,13]
           if c=='None':
               c=''
           if d=='None':
               d=''
           if e=='None':
               e=''
           
        
           
           if h== 'None'and l=='None':
               
               
              
               prueba = Programa.objects.get_or_create(Tramo = a, Dfinal= b ,de1=c , de2=d ,de3= e,d1distancia= f,Ntramo=g, Codigo=m,Darea=j,npersonas=k)
           if h!= 'None' and l == 'None':
               
               prueba = Programa.objects.get_or_create(Tramo = a, Dfinal= b ,de1=c , de2=d ,de3= e,d1distancia= f,Ntramo=g, tipoarea_id=h,Codigo=m,Darea=j,npersonas=k)
               
          
           elif h== 'None'and l!= 'None':
               
               
               prueba = Programa.objects.get_or_create(Tramo = a, Dfinal= b ,de1=c , de2=d ,de3= e,d1distancia= f,Ntramo=g,Codigo=m,Darea=j,npersonas=k,Unión_id=l)
               
               
           elif h!='None'and l!='None':
               
               
               prueba = Programa.objects.get_or_create(Tramo = a, Dfinal= b ,de1=c , de2=d ,de3= e,d1distancia= f,Ntramo=g,tipoarea_id=h,Codigo=m,Darea=j,npersonas=k,Unión_id=l)
               
               
               
    return render(request,'ventilcion/cargar.html')

def indice(request):
    return render(request,'ventilcion/index.html')



def download_file(request):  
    # do something  
   
    file_path = 'Tutorial/Tutorial.pdf' # Ruta del archivo para descargar  
    
    with open(file_path, 'rb') as f:
        try:
            response = HttpResponse(f)
            response['content_type'] = "application/octet-stream"
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
            return response
        except Exception:
            raise Http404
      


